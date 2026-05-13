# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------
import csv
import html
import json
import logging
import multiprocessing as mp
import random
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

import cherrypy
from cherrypy import _cperror

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl

import secure

from sflib import SpiderFoot

from sfscan import startSpiderFootScanner

from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup
from spiderfoot.graph_engine import SpiderFootGraphEngine
from spiderfoot.graph_correlation import GraphCorrelator

mp.set_start_method("spawn", force=True)


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:")
            .style_src("'self'", "'unsafe-inline'")
            .base_uri("'self'")
            .connect_src("'self'", "data:")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)

        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve info for scan."]).encode('utf-8')

        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve correlations for scan."]).encode('utf-8')

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.xlxs"
            else:
                fname = "SpiderFoot-correlations.xlxs"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, headings, sheetNameIndex=0)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)

            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.csv"
            else:
                fname = "SpiderFoot-correlations.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.xlsx"
            else:
                fname = scan_name + "-SpiderFoot.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.csv"
            else:
                fname = scan_name + "-SpiderFoot.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.json"
        else:
            fname = scan_name + "-SpiderFoot.json"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return SpiderFootHelpers.buildGraphJson([root], data)

        if not scan_name:
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "-SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")

            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] Scan [{scanId}] failed: {e}")
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        templ = Template(filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        """List all modules.

        Returns:
            list: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret

        modinfo.sort()

        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all correlation rules.

        Returns:
            list: list of correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        rules = self.config['__correlationrules__']
        if not rules:
            return ret

        for r in rules:
            ret.append({
                'id': r['id'],
                'name': r['meta']['name'],
                'descr': r['meta']['description'],
                'risk': r['meta']['risk'],
            })

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: SpiderFoot version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Unrecognised target type."]).encode('utf-8')

            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        # Swap the globalscantable for the database handler
        dbh = SpiderFootDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()

        # Delete the stdout module in case it crept in
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata

    #
    # SANDBOX / LINK ANALYSIS
    #

    @cherrypy.expose
    def sandbox(self: 'SpiderFootWebUi') -> str:
        """Sandbox Page.

        Returns:
            str: HTML content
        """
        templ = Template(filename='spiderfoot/templates/sandbox.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def sandboxscan(self: 'SpiderFootWebUi', url: str) -> dict:
        """Run a sandbox scan.

        Args:
            url (str): URL to scan

        Returns:
            dict: Scan results
        """
        import subprocess
        import uuid
        import os

        if not url:
            return {"status": "error", "message": "No URL provided"}

        scan_id = str(uuid.uuid4())
        # Use absolute path for output, ensure it's writable
        output_dir = os.path.abspath(os.path.join(os.getcwd(), "sandbox_out", scan_id))
        os.makedirs(output_dir, exist_ok=True)

        # Docker command
        # Mount the output_dir to /out
        # WARNING: This assumes the user has 'docker' in path and permissions
        cmd = [
            "docker", "run", "--rm",
            "-v", f"{output_dir}:/out",
            "spiderfoot-sandbox",
            url
        ]

        runner_log_file = os.path.join(output_dir, "runner.log")
        runner_output = ""
        try:
            # 200s timeout — runner.py has a 150s internal hard limit so it always exits cleanly first
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=200)
            runner_output = result.stdout
            if result.stderr:
                runner_output += "\n[STDERR]\n" + result.stderr
            # Non-zero exit is not fatal — runner may still have produced partial results
        except subprocess.TimeoutExpired as e:
            # Grab any partial output before timeout
            runner_output = (e.stdout or "") + "\n\n[HOST] Analysis timed out after 200 seconds."
        except Exception as e:
            return {"status": "error", "message": str(e)}

        # Always save runner process log so the UI can display it
        try:
            with open(runner_log_file, "w") as f:
                f.write(runner_output)
        except Exception:
            pass

        # Check for artifacts
        # runner.py produces screenshots.json manifest + screenshot_N_name.png files
        manifest_file = os.path.join(output_dir, "screenshots.json")
        log_file = os.path.join(output_dir, "console.json")

        has_screenshot = os.path.exists(manifest_file)
        has_logs = os.path.exists(log_file)
        has_runner_log = os.path.exists(runner_log_file)

        # Read manifest to get list of screenshot files
        screenshots = []
        total_screenshots = 0
        if has_screenshot:
            try:
                with open(manifest_file, "r") as f:
                    import json as _json
                    manifest = _json.load(f)
                    screenshots = manifest.get("screenshots", [])
                    total_screenshots = manifest.get("total_screenshots", len(screenshots))
            except Exception:
                pass

        return {
            "status": "success",
            "scan_id": scan_id,
            "has_screenshot": has_screenshot,
            "has_logs": has_logs,
            "has_runner_log": has_runner_log,
            "total_screenshots": total_screenshots,
            "screenshots": screenshots
        }

    @cherrypy.expose
    def sandboxfile(self: 'SpiderFootWebUi', id: str, file: str) -> bytes:
        """Serve sandbox artifacts."""
        import os
        import re
        
        # Security check: Ensure id is valid uuid and file is allowed
        # Simple alphanum check for id to prevent traversal
        if not id or not id.replace("-", "").isalnum() or ".." in file:
             return self.error("Invalid request")

        # Allow screenshot_*.png files, console.json, screenshots.json, and runner.log
        valid_patterns = [r'^screenshot_\d+_.*\.png$', r'^console\.json$', r'^screenshots\.json$', r'^runner\.log$']
        if not any(re.match(pattern, file) for pattern in valid_patterns):
            return self.error("Invalid file type")

        path = os.path.abspath(os.path.join(os.getcwd(), "sandbox_out", id, file))
        
        if not os.path.exists(path):
            return self.error("File not found")

        if file.endswith(".png"):
            cherrypy.response.headers['Content-Type'] = "image/png"
        elif file.endswith(".json"):
            cherrypy.response.headers['Content-Type'] = "application/json"
        elif file.endswith(".log"):
            cherrypy.response.headers['Content-Type'] = "text/plain; charset=utf-8"

        with open(path, "rb") as f:
            return f.read()

    #
    # GRAPH OPERATIONS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scangraphbuild(self: 'SpiderFootWebUi', id: str) -> dict:
        """Build graph for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: Status
        """
        engine = SpiderFootGraphEngine(self.config)
        graph = engine.build_graph_from_scan(id)
        
        if graph:
            return {"status": "SUCCESS", "nodes": graph.number_of_nodes(), "edges": graph.number_of_edges()}
        else:
            return {"status": "ERROR", "message": "Graph build failed"}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scangraphanalyze(self: 'SpiderFootWebUi', id: str) -> dict:
        """Run graph analysis.

        Args:
            id (str): scan ID

        Returns:
            dict: Status
        """
        correlator = GraphCorrelator(self.config, id)
        correlator.run()
        return {"status": "SUCCESS", "message": "Graph correlation complete"}

    @cherrypy.expose
    def scangraphexport(self: 'SpiderFootWebUi', id: str, format: str = "gexf") -> bytes:
        """Export graph.

        Args:
             id (str): scan ID
             format (str): export format
        
        Returns:
             bytes: file content
        """
        engine = SpiderFootGraphEngine(self.config)
        graph = engine.build_graph_from_scan(id)
        
        if not graph:
             return self.error("Graph build failed")
             
        if format == "gexf":
            from io import BytesIO
            import networkx as nx
            
            f = BytesIO()
            nx.write_gexf(graph, f)
            
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.gexf"
            cherrypy.response.headers['Content-Type'] = "application/gexf"
            return f.getvalue()
            
        elif format == "json":
             import json
             import networkx as nx
             data = nx.node_link_data(graph)
             cherrypy.response.headers['Content-Type'] = "application/json"
             return json.dumps(data).encode('utf-8')
             
        return self.error("Invalid format")

    #
    # RISK SCORING
    #

    @cherrypy.expose
    def riskanalysis(self):
        """Entity risk analysis page."""
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        templ = Template(filename='spiderfoot/templates/riskanalysis.tmpl', lookup=self.lookup)
        return templ.render(scans=scans, docroot=self.docroot, version=__version__, pageid="RISK")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def riskscore(self, scan_id=None):
        """Run risk scoring engine on a scan and return results as JSON."""
        if not scan_id:
            return {"status": "error", "message": "No scan ID specified"}

        try:
            import time as _time
            from spiderfoot.risk_scorer import EntityRiskScorer

            t0 = _time.time()
            dbh = SpiderFootDb(self.config)

            scorer = EntityRiskScorer(dbh)
            scored = scorer.score_all_entities(scan_id)
            summary = scorer.get_scan_risk_summary(scored)

            # Persist to DB (delete old scores first for idempotency)
            dbh.deleteRiskScores(scan_id)
            scorer.store_risk_scores(scan_id, scored)

            elapsed = round(_time.time() - t0, 2)

            return {
                "status": "success",
                "elapsed": elapsed,
                "summary": summary,
                "entities": scored,
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def riskscoredata(self, scan_id=None, level=None):
        """Retrieve stored risk scores for a scan."""
        if not scan_id:
            return {"status": "error", "message": "No scan ID specified"}

        try:
            dbh = SpiderFootDb(self.config)
            rows = dbh.getRiskScores(scan_id, riskLevel=level)

            entities = []
            for r in rows:
                entities.append({
                    "id": r[0],
                    "entity_hash": r[2],
                    "entity_type": r[3],
                    "entity_data": r[4],
                    "risk_score": r[5],
                    "risk_level": r[6],
                    "threat_score": r[7],
                    "graph_score": r[8],
                    "vuln_score": r[9],
                    "infra_score": r[10],
                    "exposure_score": r[11],
                    "details": r[12],
                    "created_time": r[13],
                })

            return {"status": "success", "entities": entities}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @cherrypy.expose
    def graphanalysis(self):
        """Graph-based correlation analysis page"""
        # Get all completed scans
        dbh = SpiderFootDb(self.config)
        scans = dbh.scanInstanceList()
        
        templ = Template(filename='spiderfoot/templates/graphanalysis.tmpl', lookup=self.lookup)
        return templ.render(scans=scans, docroot=self.docroot, version=__version__, pageid="GRAPH")
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def graphbuild(self, scan_id):
        """Build graph for a scan"""
        try:
            from spiderfoot.graph_engine import SpiderFootGraphEngine
            
            dbh = SpiderFootDb(self.config)
            
            # Initialize engine
            engine = SpiderFootGraphEngine(dbh, scan_id)
            
            # Build graph
            graph = engine.build_graph_from_scan(include_affiliates=True)
            
            # Get statistics
            stats = engine.get_graph_stats()
            
            # Save graph to disk for later use
            import os
            graph_dir = f"/tmp/prism_graphs"
            os.makedirs(graph_dir, exist_ok=True)
            engine.save_to_disk(f"{graph_dir}/{scan_id}.pkl")
            
            return {
                'status': 'success',
                'stats': stats
            }
        except Exception as e:
            return {
                'status': 'error',
                'message': str(e)
            }
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def graphcorrelations(self, scan_id):
        """Get graph correlations for a scan"""
        try:
            from spiderfoot.graph_engine import SpiderFootGraphEngine
            from spiderfoot.graph_algorithms import GraphAnalyzer
            import os
            import json as _json

            dbh = SpiderFootDb(self.config)

            # Load or build graph
            graph_path = f"/tmp/prism_graphs/{scan_id}.pkl"
            engine = SpiderFootGraphEngine(dbh, scan_id)

            if os.path.exists(graph_path):
                engine.load_from_disk(graph_path)
            else:
                engine.build_graph_from_scan()

            # Run in-memory analysis
            analyzer = GraphAnalyzer(engine.graph)
            correlations_data = analyzer.run_all_correlations()
            summary = correlations_data.get("summary", {})

            # Delete old entries for re-run freshness
            with dbh.dbhLock:
                dbh.dbh.execute(
                    "DELETE FROM tbl_scan_graph_correlations WHERE scan_instance_id = ?",
                    (scan_id,))
                dbh.conn.commit()

            # Enrich each result type with metadata and store
            store_qry = """INSERT INTO tbl_scan_graph_correlations
                (scan_instance_id, correlation_type, algorithm, entities, metadata, score, created_time)
                VALUES (?, ?, ?, ?, ?, ?, datetime('now'))"""

            communities = correlations_data.get("communities", [])
            centrality  = correlations_data.get("centrality", [])
            anomalies   = correlations_data.get("anomalies", [])

            # Build metadata dicts with meaningful descriptions
            num_c = len(communities)
            large = [c for c in communities if c.get("size", 0) > 5]
            comm_risk = "HIGH" if large else ("MEDIUM" if num_c > 3 else "LOW")
            comm_meta = {
                "title": f"{num_c} Entity Communities Detected",
                "description": (
                    f"Graph clustering found {num_c} distinct communities. "
                    f"{len(large)} large cluster(s) (>5 nodes) suggest potential "
                    f"coordinated infrastructure or threat-actor grouping. "
                    f"Largest group has {communities[0]['size'] if communities else 0} entities."
                ),
                "risk_level": comm_risk,
            }

            high_pr = [n for n in centrality if n.get("risk_level") == "HIGH"]
            cent_risk = "HIGH" if high_pr else ("MEDIUM" if centrality else "LOW")
            cent_meta = {
                "title": f"{len(high_pr)} High-Centrality Pivot Nodes Found",
                "description": (
                    f"PageRank analysis identified {len(centrality)} nodes by influence. "
                    f"{len(high_pr)} node(s) scored in the HIGH tier — these act as "
                    f"infrastructure bridges and should be prioritised in attribution."
                    + (f" Top pivot: {high_pr[0]['label']} ({high_pr[0]['type']})" if high_pr else "")
                ),
                "risk_level": cent_risk,
            }

            high_anom = [a for a in anomalies if a.get("risk_level") == "HIGH"]
            anom_risk = "HIGH" if high_anom else ("MEDIUM" if anomalies else "LOW")
            anom_meta = {
                "title": f"{len(anomalies)} Structural Anomalies ({len(high_anom)} HIGH)",
                "description": (
                    f"Heuristic analysis flagged {len(anomalies)} graph anomalies. "
                    f"{len(high_anom)} high-degree node(s) with abnormally many connections "
                    f"may indicate command-and-control infrastructure or data aggregation hubs. "
                    + (f"Most suspect: {high_anom[0]['node']} (degree={high_anom[0]['degree']})" if high_anom else "No high-risk anomalies.")
                ),
                "risk_level": anom_risk,
            }

            risk_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
            for corr_type, algo, dataset, meta in [
                ("community", "louvain",   communities, comm_meta),
                ("centrality","pagerank",  centrality,  cent_meta),
                ("anomaly",   "heuristic", anomalies,   anom_meta),
            ]:
                score = risk_order.get(meta["risk_level"], 1) / 3.0
                try:
                    with dbh.dbhLock:
                        dbh.dbh.execute(store_qry, (
                            scan_id, corr_type, algo,
                            _json.dumps(dataset[:20]),   # cap payload
                            _json.dumps(meta),
                            score,
                        ))
                        dbh.conn.commit()
                except Exception:
                    pass

            db_correlations = dbh.getGraphCorrelations(scan_id)
            return {'status': 'success', 'correlations': db_correlations}

        except Exception as e:
            return {'status': 'error', 'message': str(e), 'correlations': []}



    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def graphdata(self, scan_id):
        """Get graph data for visualization"""
        try:
            from spiderfoot.graph_engine import SpiderFootGraphEngine
            import os
            
            dbh = SpiderFootDb(self.config)
            
            # Load graph
            graph_path = f"/tmp/prism_graphs/{scan_id}.pkl"
            
            engine = SpiderFootGraphEngine(dbh, scan_id)
            
            if os.path.exists(graph_path):
                engine.load_from_disk(graph_path)
            else:
                engine.build_graph_from_scan()
            
            # Convert to D3.js format
            nodes = []
            links = []
            
            for node_id, attrs in engine.graph.nodes(data=True):
                nodes.append({
                    'id': node_id,
                    'label': attrs.get('label', attrs.get('data', node_id))[:30],
                    'type': attrs.get('type', 'UNKNOWN'),
                    'data': attrs.get('data', ''),
                    'color': attrs.get('color', '#888')
                })
            
            for source, target, attrs in engine.graph.edges(data=True):
                links.append({
                    'source': source,
                    'target': target,
                    'type': attrs.get('type', 'RELATED')
                })
            
            return {
                'status': 'success',
                'graph': {
                    'nodes': nodes,
                    'links': links
                }
            }
        except Exception as e:
            return {
                'status': 'error',
                'message': str(e)
            }
    
    @cherrypy.expose
    def graphexport(self, scan_id, format='json'):
        """Export graph in various formats"""
        try:
            from spiderfoot.graph_engine import SpiderFootGraphEngine
            import os
            
            dbh = SpiderFootDb(self.config)
            
            # Load graph
            graph_path = f"/tmp/prism_graphs/{scan_id}.pkl"
            engine = SpiderFootGraphEngine(dbh, scan_id)
            
            if os.path.exists(graph_path):
                engine.load_from_disk(graph_path)
            else:
                engine.build_graph_from_scan()
            
            # Export
            export_path = engine.export_graph(format=format, filepath=f"/tmp/graph_{scan_id}.{format}")
            
            # Read and return file
            with open(export_path, 'rb') as f:
                data = f.read()
            
            cherrypy.response.headers['Content-Type'] = 'application/octet-stream'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="graph_{scan_id}.{format}"'
            
            return data
        except Exception as e:
            return f"Error: {str(e)}"
